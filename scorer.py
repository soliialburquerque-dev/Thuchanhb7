import os
import csv
import json
import re
import io
import urllib.request
import urllib.parse

# ─── File Paths ───────────────────────────────────────────────────────────────
BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
CSV_PATH         = os.path.join(BASE_DIR, "data.csv")
JSON_PATH        = os.path.join(BASE_DIR, "leads_processed.json")
CRITERIA_PATH    = os.path.join(BASE_DIR, "tieu_chi_cham_diem.txt")
CREDENTIALS_PATH = os.path.join(BASE_DIR, "credentials.json")

# Google Sheet config
SPREADSHEET_ID = "16tCAf_qqtgYZxoumYQKMEOdBhKE0wg5A"
SHEET_GID      = "1542775777"
SHEET_CSV_URL  = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=csv&gid={SHEET_GID}"


# ─── Google Sheets Fetcher ────────────────────────────────────────────────────
def fetch_sheet_data():
    """
    Download all rows from the Google Sheet.
    Priority:
      1. Service Account via gspread (native Google Sheets) or Drive download (Excel)
      2. Public CSV export URL (fallback)
      3. Local data.csv cache (offline fallback)
    Returns list of dicts with keys: id, ten_khach, sdt, nhu_cau_mo_ta
    """
    # --- Method 1: Service Account ---
    if os.path.exists(CREDENTIALS_PATH):
        try:
            from google.oauth2.service_account import Credentials
            scopes = [
                "https://www.googleapis.com/auth/drive.readonly",
                "https://www.googleapis.com/auth/spreadsheets.readonly",
            ]
            creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=scopes)

            # Try 1a: gspread (native Google Sheets)
            try:
                import gspread
                gc = gspread.authorize(creds)
                sh = gc.open_by_key(SPREADSHEET_ID)
                target_gid = int(SHEET_GID)
                ws = next((s for s in sh.worksheets() if s.id == target_gid), sh.get_worksheet(0))
                records = ws.get_all_records()
                print(f"[gspread] Loaded {len(records)} rows via Sheets API.")
                _save_cache(records)
                return records
            except Exception as e1:
                print(f"[gspread] {e1}. Trying Drive download...")

            # Try 1b: Drive API download raw file (works for Excel .xlsx)
            from googleapiclient.discovery import build
            drive = build("drive", "v3", credentials=creds)
            # Get file metadata to determine MIME type
            meta = drive.files().get(fileId=SPREADSHEET_ID, fields="mimeType,name").execute()
            mime = meta.get("mimeType", "")
            name = meta.get("name", "")
            print(f"[drive] File: {name!r}, mimeType: {mime}")

            if "spreadsheet" in mime:
                # Native Google Sheets — export as CSV
                req = drive.files().export_media(fileId=SPREADSHEET_ID, mimeType="text/csv")
            else:
                # Excel or other — download raw bytes
                req = drive.files().get_media(fileId=SPREADSHEET_ID)

            buf = io.BytesIO()
            from googleapiclient.http import MediaIoBaseDownload
            dl = MediaIoBaseDownload(buf, req)
            done = False
            while not done:
                _, done = dl.next_chunk()
            raw = buf.getvalue()

            if "spreadsheet" in mime:
                # CSV content
                records = list(csv.DictReader(io.StringIO(raw.decode("utf-8"))))
                print(f"[drive-export] Loaded {len(records)} rows as CSV.")
            else:
                # Excel content — parse with openpyxl
                try:
                    import openpyxl
                except ImportError:
                    import subprocess, sys
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
                    import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                # Find correct sheet by index matching GID order (best effort)
                ws_names = wb.sheetnames
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                records = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    records.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
                wb.close()
                print(f"[drive-xlsx] Loaded {len(records)} rows from Excel file.")

            _save_cache(records)
            return records

        except Exception as e:
            print(f"[service-account] Failed: {e}. Trying public URL...")

    # --- Method 2: Public CSV URL ---
    try:
        req = urllib.request.Request(SHEET_CSV_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            content = resp.read().decode("utf-8")
        records = list(csv.DictReader(io.StringIO(content)))
        print(f"[url] Loaded {len(records)} rows from public CSV URL.")
        with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
            f.write(content)
        return records
    except Exception as e:
        print(f"[url] Failed: {e}. Using local cache...")

    # --- Method 3: Local CSV cache ---
    if os.path.exists(CSV_PATH):
        with open(CSV_PATH, "r", encoding="utf-8") as f:
            records = list(csv.DictReader(f))
        print(f"[local] Loaded {len(records)} rows from cache.")
        return records

    print("[ERROR] No data source available!")
    return []


def _save_cache(records):
    """Save records list to local CSV cache."""
    if not records:
        return
    keys = list(records[0].keys())
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(records)
    print(f"[cache] Saved {len(records)} rows to {CSV_PATH}")




# ─── Criteria Reader ─────────────────────────────────────────────────────────
def read_criteria():
    if os.path.exists(CRITERIA_PATH):
        with open(CRITERIA_PATH, "r", encoding="utf-8") as f:
            return f.read()
    return "Rules not found. Please refer to tieu_chi_cham_diem.txt."


# ─── Rule-Based Scorer ───────────────────────────────────────────────────────
def rule_based_score(text):
    """
    Evaluates score based on local keyword matching from tieu_chi_cham_diem.txt.
    Base score: 50.
    Adds 50 for VIP indicators, subtracts 50 for spam/unqualified indicators.
    """
    text_lower = text.lower()
    score = 50
    reasons = []

    # VIP Criteria Indicators (+50)
    vip_matched = False

    # Budget check (>= 20B or keyword phrases)
    budget_keywords = ["tài chính mạnh", "không thành vấn đề", "tài chính cực mạnh", "ngân sách cực mạnh"]
    has_large_budget = any(kw in text_lower for kw in budget_keywords)
    budget_nums = re.findall(r'(\d+)\s*(?:tỷ|tỉ)', text_lower)
    for num in budget_nums:
        if int(num) >= 20:
            has_large_budget = True
            break
    if has_large_budget:
        reasons.append("+50: Ngân sách lớn (>= 20 tỷ hoặc tài chính mạnh/không thành vấn đề)")
        vip_matched = True

    # Luxury property check
    luxury_keywords = ["biệt thự đơn lập", "penthouse", "shophouse mặt đường lớn",
                       "quỹ đất công nghiệp", "sàn văn phòng diện tích lớn",
                       "shophouse", "sản văn phòng", "đất công nghiệp"]
    if any(kw in text_lower for kw in luxury_keywords):
        reasons.append("+50: Loại hình cao cấp (Biệt thự, Penthouse, Shophouse, Đất công nghiệp/Văn phòng)")
        vip_matched = True

    # Prime location check
    location_keywords = ["quận 1", "ven sông", "vinhomes ocean park", "phú mỹ hưng", "q1", "q.1"]
    if any(kw in text_lower for kw in location_keywords):
        reasons.append("+50: Vị trí đắc địa (Quận 1, Ven sông, Vinhomes Ocean Park, Phú Mỹ Hưng)")
        vip_matched = True

    # Target profile check
    profile_keywords = ["chủ doanh nghiệp", "nhà đầu tư chuyên nghiệp", "mua sỉ", "mua số lượng lớn", "gom sỉ"]
    if any(kw in text_lower for kw in profile_keywords):
        reasons.append("+50: Đối tượng khách hàng VIP (Chủ DN, Mua sỉ, Đầu tư chuyên nghiệp)")
        vip_matched = True

    # Urgency & Transparency check
    urgency_keywords = ["pháp lý chuẩn 100%", "sổ hồng riêng", "gặp trực tiếp chủ đầu tư",
                        "gặp trực tiếp giám đốc", "gặp trực tiếp chủ đầu tư để đàm phán"]
    if any(kw in text_lower for kw in urgency_keywords):
        reasons.append("+50: Tính cấp thiết & Minh bạch cao (Sổ hồng riêng, Pháp lý 100%, Đàm phán trực tiếp)")
        vip_matched = True

    if vip_matched:
        score += 50

    # Spam/Unqualified Criteria Indicators (-50)
    spam_matched = False

    # Unrealistic requirements
    unrealistic_keywords = ["giá thấp vô lý", "q1 giá 1 tỷ", "quận 1 giá 1",
                            "quận 1 giá 2 tỷ", "thuê nguyên căn giá 2 triệu",
                            "yêu cầu phi thực tế", "ngân sách rất thấp"]
    has_unrealistic = any(kw in text_lower for kw in unrealistic_keywords) or \
                      ("q1" in text_lower and "1 tỷ" in text_lower)
    if has_unrealistic:
        reasons.append("-50: Yêu cầu phi thực tế (Giá quá rẻ so với thị trường)")
        spam_matched = True

    # No intent
    no_intent_keywords = ["nhầm số", "không có nhu cầu", "dữ liệu cũ", "nhầm ngành"]
    if any(kw in text_lower for kw in no_intent_keywords):
        reasons.append("-50: Không có nhu cầu (Nhầm số, dữ liệu cũ, nhầm ngành)")
        spam_matched = True

    # Uncooperative
    uncooperative_keywords = ["hỏi giá cho vui", "chưa có ý định mua", "thái độ không hợp tác", "không hợp tác"]
    if any(kw in text_lower for kw in uncooperative_keywords):
        reasons.append("-50: Khách hàng không thiện chí (Hỏi vui, chưa muốn mua, thái độ kém)")
        spam_matched = True

    # Spam/Advertising
    spam_keywords = ["quảng cáo ngược", "bảo hiểm", "vay vốn", "mời chào", "chào mời",
                     "quảng cáo ngược lại dịch vụ"]
    has_spam = any(kw in text_lower for kw in spam_keywords) and \
               not ("cần hỗ trợ vay" in text_lower or "cần vay" in text_lower)
    if has_spam or "spam" in text_lower:
        reasons.append("-50: Tin nhắn Spam / Quảng cáo dịch vụ")
        spam_matched = True

    # Contact error
    contact_keywords = ["thuê bao", "gọi nhiều lần không bắt máy", "không phản hồi zalo", "gọi không liên lạc được"]
    if any(kw in text_lower for kw in contact_keywords):
        reasons.append("-50: Thông tin liên lạc lỗi (Thuê bao, không bắt máy, không trả lời Zalo)")
        spam_matched = True

    if spam_matched:
        score -= 50

    # Normal / Mid-range cases
    if not vip_matched and not spam_matched:
        normal_keywords = ["chung cư", "nhà phố", "căn hộ 2pn", "căn hộ 3pn", "tầm trung",
                           "3-10 tỷ", "vay ngân hàng", "tư vấn thêm", "nhà phố liền kề"]
        if any(kw in text_lower for kw in normal_keywords):
            reasons.append("+10: Khách hàng nhu cầu thực phân khúc trung cấp (Căn hộ, Nhà phố 3-10 tỷ, Cần vay ngân hàng)")
            score = 60
        else:
            reasons.append("0: Khách hàng bình thường / Cần tư vấn thêm")
            score = 50

    score = max(0, min(100, score))

    if score >= 80:
        status = "VIP"
    elif score < 30:
        status = "SPAM"
    else:
        status = "NORMAL"

    return score, "; ".join(reasons), status


# ─── Gemini AI Scorer ────────────────────────────────────────────────────────
def score_with_gemini(text, api_key):
    """Calls Gemini API to score the lead based on tieu_chi_cham_diem.txt."""
    criteria = read_criteria()
    prompt = f"""
Bạn là chuyên gia phân loại và chấm điểm khách hàng tiềm năng (Lead Scoring AI) cho ngành Bất động sản.
Nhiệm vụ của bạn là đọc mô tả nhu cầu của khách hàng, phân tích và chấm điểm dựa trên bộ quy tắc sau:

{criteria}

Hãy chấm điểm theo thang điểm từ 0 đến 100:
- Điểm mặc định bắt đầu: 50 điểm.
- Cộng thêm 50 điểm nếu có ít nhất một tiêu chí VIP/SIÊU TIỀM NĂNG.
- Trừ đi 50 điểm nếu có ít nhất một tiêu chí RÁC/SPAM/KHÔNG TIỀM NĂNG.
- Phân khúc trung cấp (chung cư, nhà phố 3-10 tỷ, cần vay ngân hàng): 50-60 điểm.

Phân loại:
- "VIP" nếu Điểm >= 80
- "SPAM" nếu Điểm < 30
- "NORMAL" nếu Điểm từ 30 đến 79

Trả về JSON duy nhất:
{{
  "score": <số nguyên 0-100>,
  "status": "<VIP, NORMAL hoặc SPAM>",
  "reason": "<lý do chi tiết>"
}}

Nhu cầu khách hàng: "{text}"
"""
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    data = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json"}
    }
    req_body = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=req_body, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as response:
            res_body = response.read().decode("utf-8")
            res_json = json.loads(res_body)
            text_response = res_json['candidates'][0]['content']['parts'][0]['text'].strip()
            if text_response.startswith("```json"):
                text_response = text_response[7:]
            if text_response.endswith("```"):
                text_response = text_response[:-3]
            parsed = json.loads(text_response.strip())
            score  = max(0, min(100, int(parsed.get("score", 50))))
            status = parsed.get("status", "NORMAL").upper()
            reason = parsed.get("reason", "Phân tích bằng AI.")
            if status not in ["VIP", "NORMAL", "SPAM"]:
                status = "VIP" if score >= 80 else ("SPAM" if score < 30 else "NORMAL")
            return score, f"[AI] {reason}", status
    except Exception as e:
        print(f"Gemini API request failed: {e}. Falling back to Rule-Based Scorer.")
        return rule_based_score(text)


# ─── Main Scoring Pipeline ────────────────────────────────────────────────────
def run_scoring(api_key=None, overwrite=False, fetch_from_sheet=True):
    """
    Fetch data from Google Sheet (via Service Account or public URL),
    score all leads, merge with existing manual edits, and save to JSON.
    """
    # Load existing processed data to preserve manual edits
    existing_data = {}
    if not overwrite and os.path.exists(JSON_PATH):
        try:
            with open(JSON_PATH, "r", encoding="utf-8") as f:
                leads_list = json.load(f)
                existing_data = {str(item["id"]): item for item in leads_list}
        except Exception as e:
            print(f"Error loading existing leads: {e}")

    # Fetch rows
    if fetch_from_sheet:
        rows = fetch_sheet_data()
    elif os.path.exists(CSV_PATH):
        with open(CSV_PATH, "r", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
    else:
        print("No data source available!")
        return False

    if not rows:
        print("No rows fetched.")
        return False

    processed_leads = []
    for row in rows:
        lead_id = str(row["id"])
        name    = str(row["ten_khach"])
        phone   = str(row["sdt"])
        desc    = str(row["nhu_cau_mo_ta"])

        # Keep manually reviewed leads as-is
        if lead_id in existing_data and existing_data[lead_id].get("reviewed", False):
            processed_leads.append(existing_data[lead_id])
            continue

        # Score
        if api_key:
            score, reason, status = score_with_gemini(desc, api_key)
        else:
            score, reason, status = rule_based_score(desc)

        processed_leads.append({
            "id":            int(lead_id),
            "ten_khach":     name,
            "sdt":           phone,
            "nhu_cau_mo_ta": desc,
            "score":         score,
            "reason":        reason,
            "status":        status,
            "reviewed":      False,
            "notes":         existing_data.get(lead_id, {}).get("notes", ""),
        })

    processed_leads.sort(key=lambda x: x["id"])

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(processed_leads, f, ensure_ascii=False, indent=2)

    print(f"Successfully processed and saved {len(processed_leads)} leads.")
    return True


# Alias for backward compatibility
calculate_score = rule_based_score

if __name__ == "__main__":
    run_scoring()
